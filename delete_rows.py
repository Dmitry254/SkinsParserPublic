import openpyxl
import traceback
from datetime import datetime, timedelta
from json import JSONDecodeError


def delete_rows():
    sheet = excel_items_table['items']
    new_row = sheet.max_row + 1
    for count in range(sheet.max_row):
        if item_info[2] == sheet.cell(row=count + 1, column=4).value and item_info[0] == sheet.cell(row=count + 1,
                                                                                                    column=1).value:
            start = False
            break
    if start:
        sheet.cell(row=new_row, column=1).value = item_info[0]
        sheet.cell(row=new_row, column=2).value = item_info[1]
        sheet.cell(row=new_row, column=4).value = item_info[2]
        if int(bot_number) in need_bots_numbers:
            sheet.cell(row=new_row, column=6).value = bot_number


if __name__ == "__main__":
    excel_items_table = openpyxl.load_workbook(filename="test.xlsx")
    try:
        delete_rows()
    finally:
        excel_items_table.save('test.xlsx')