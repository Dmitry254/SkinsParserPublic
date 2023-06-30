import openpyxl


def set_yes():
    sheet = excel_items_table['items']
    for count in range(sheet.max_row):
        if sheet.cell(row=count + 1, column=3).value is not None:
            sheet.cell(row=count + 1, column=9).value = "да"


if __name__ == "__main__":
    excel_items_table = openpyxl.load_workbook(filename="items_table.xlsx")
    try:
        set_yes()
    finally:
        excel_items_table.save('items_table.xlsx')

