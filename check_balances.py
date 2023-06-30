import openpyxl
import json
import requests


def check_balances():
    sheet = excel_items_table['Аккаунты']
    for count in range(sheet.max_row):
        api_key = sheet.cell(row=count + 1, column=4).value
        url = f"https://market.csgo.com/api/v2/get-money?key={api_key}"
        result = requests.get(url)
        my_balance = json.loads(result.text)
        print(f"{count}. Ключ {api_key}, баланс {my_balance}")


def check_balance(api_key):
        url = f"https://market.csgo.com/api/v2/get-money?key={api_key}"
        result = requests.get(url)
        my_balance = json.loads(result.text)
        print(my_balance)


if __name__ == "__main__":
    check_balance("")
    # excel_items_table = openpyxl.load_workbook(filename="items_table.xlsx")
    # try:
    #     check_balances()
    # finally:
    #     excel_items_table.save('items_table.xlsx')
