import openpyxl
import json
import requests
import traceback
from datetime import datetime, timedelta
from json import JSONDecodeError
from key import *


def get_history(date):
    time_end = datetime.now().timestamp()
    url = f"https://market.csgo.com/api/v2/history?key={api_key}&date={date}&date_end={time_end}"
    result = requests.get(url)
    data = json.loads(result.text)
    for item in reversed(range(len(data['data']))):
        if data['data'][item]['event'] == "buy" and data['data'][item]['stage'] == "2":
            buy_items_table(excel_items_table, data['data'][item]['market_hash_name'],
                            float(data['data'][item]['paid']) / 100, data['data'][item]['item_id'])
        if data['data'][item]['event'] == "sell" and data['data'][item]['stage'] == "2":
            sold_items_table(excel_items_table, data['data'][item]['market_hash_name'],
                             float(data['data'][item]['received']) / 100, data['data'][item]['item_id'])


def buy_items_table(excel_items_table, *item_info):
    start = True
    sheet = excel_items_table['items']
    new_row = sheet.max_row + 1
    for count in range(sheet.max_row):
        if item_info[2] == sheet.cell(row=count + 1, column=4).value and item_info[0] == sheet.cell(row=count + 1,
                                                                                                    column=1).value:
            start = False
            break
    if start:
        sheet.cell(row=new_row, column=1).value = item_info[0]
        sheet.cell(row=new_row, column=2).value = item_info[1]
        sheet.cell(row=new_row, column=4).value = item_info[2]
        if int(bot_number) in need_bots_numbers:
            sheet.cell(row=new_row, column=6).value = bot_number


def sold_items_table(excel_items_table, *item_info):
    start = True
    sheet = excel_items_table['items']
    for count in range(sheet.max_row):
        if item_info[2] == sheet.cell(row=count + 1, column=5).value and item_info[0] == sheet.cell(row=count + 1,
                                                                                                    column=1).value:
            start = False
            break
    if start:
        for counter in range(sheet.max_row):
            if item_info[0] == sheet.cell(row=counter + 1, column=1).value \
                    and sheet.cell(row=counter + 1, column=3).value is None and sheet.cell(row=counter + 1,
                                                                                           column=5).value is None:
                sheet.cell(row=counter + 1, column=3).value = item_info[1]
                sheet.cell(row=counter + 1, column=5).value = item_info[2]
                break


def count_trades():
    buy_count = 0
    sell_count = 0
    start_date = datetime(current_year, current_month, current_day).timestamp()
    url = f"https://market.csgo.com/api/v2/history?key={api_key}&date={start_date}"
    result = requests.get(url)
    data = json.loads(result.text)
    for item in reversed(range(len(data['data']))):
        if data['data'][item]['event'] == "buy" and data['data'][item]['stage'] == "2":
            buy_count += 1
        if data['data'][item]['event'] == "sell" and data['data'][item]['stage'] == "2":
            sell_count += 1
    print(f"Бот {bot_number}: покупок {buy_count}, продаж {sell_count}")


def get_balances():
    count_bots = 0
    total_bots = len(api_keys)
    try:
        text = ""
        while count_bots != total_bots:
            try:
                key = api_keys[count_bots]
                item_count = 0
                item_sum = 0
                url = f"https://market.csgo.com/api/v2/get-money?key={key}"
                api = requests.get(url)
                data = json.loads(api.text)
                url1 = f"https://market.csgo.com/api/v2/items?key={key}"
                api1 = requests.get(url1)
                data1 = json.loads(api1.text)
                if data1['items']:
                    for item in data1['items']:
                        if item['status'] == "1":
                            item_count += 1
                            item_sum += item['price'] * 0.95
                item_sum = int(item_sum)
                text += f"На {api_keys.index(key) + 1} боте {data['money']} и продаётся {item_count} вещей на {item_sum}\n"
                count_bots += 1
            except JSONDecodeError:
                pass
        return text
    except:
        traceback.print_exc()
        return "Неизвестная ошибка"


if __name__ == "__main__":
    print(get_balances())
    count_bots = 0
    total_bots = len(api_keys)
    need_bots_numbers = [1, 2]
    days_list = ["Вчера", "Сегодня"]
    date = (datetime.today() - timedelta(days=3)).timestamp()
    excel_items_table = openpyxl.load_workbook(filename="items_table.xlsx")
    try:
        for day_info in days_list:
            print(day_info)
            current_day = int((datetime.today() + timedelta(days=days_list.index(day_info) - 1)).strftime("%d"))
            current_month = int((datetime.today() + timedelta(days=days_list.index(day_info) - 1)).strftime("%m"))
            current_year = int((datetime.today() + timedelta(days=days_list.index(day_info) - 1)).strftime("%Y"))
            while count_bots != total_bots:
                try:
                    api_key = api_keys[count_bots]
                    bot_number = api_keys.index(api_key) + 1
                    get_history(date)
                    count_trades()
                    count_bots += 1
                except JSONDecodeError:
                    pass
            count_bots = 0
    finally:
        excel_items_table.save('items_table.xlsx')
